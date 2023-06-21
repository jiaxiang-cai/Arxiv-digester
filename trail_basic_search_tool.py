import arxiv
import json
import openpyxl
from openpyxl.utils import get_column_letter

#这个文件就是把basic_search_tool 里面的def部分，和不同function里面叫其他function部分去掉，用quantum作为例子看看是否能下载保存应有的文档到json和excel里面去
# first part search file 
search = arxiv.Search(
    query="quantum",
    max_results=10,
    sort_by=arxiv.SortCriterion.SubmittedDate
)

# Saving search results to JSON
results = []
for result in search.results():
    authors = [str(author) for author in result.authors]  # Convert Author object to string
    result_info = {
        'title': result.title,
        'authors': authors,
        'abstract': result.summary,
        'upload_date': result.updated.date().isoformat()
    }
    results.append(result_info)

with open('search_results.json', 'w') as json_file:
    json.dump(results, json_file, indent=4)

# Saving search results to Excel
wb = openpyxl.Workbook()
ws = wb.active
header = ['Title', 'Authors', 'Abstract', 'Upload Date']
for col_num, column_title in enumerate(header, 1):
    col_letter = get_column_letter(col_num)
    ws[f"{col_letter}1"] = column_title

for row_num, result in enumerate(results, 2):
    ws[f"A{row_num}"] = result['title']
    ws[f"B{row_num}"] = ", ".join(result['authors'])
    ws[f"C{row_num}"] = result['abstract']
    ws[f"D{row_num}"] = result['upload_date']

wb.save('search_results.xlsx')



# download search results
import requests

num_download = 5  # Specify the number of papers to download

results = search.results()
downloaded_papers = []
for i in range(num_download):
    try:
        result = next(results)
        pdf_url = result.pdf_url
        response = requests.get(pdf_url)
        downloaded_papers.append(response.content)
        print(f"Downloaded paper {i+1}/{num_download}")
    except StopIteration:
        print("Reached the end of search results.")
        break
    except Exception as e:
        print(f"Error occurred while downloading paper {i+1}: {str(e)}")

# Save downloaded papers as PDF files
for i, paper in enumerate(downloaded_papers):
    try:
        with open(f"paper_{i+1}.pdf", "wb") as file:
            file.write(paper)
        print(f"Saved paper {i+1} as PDF file.")
    except Exception as e:
        print(f"Error occurred while saving paper {i+1}: {str(e)}")


