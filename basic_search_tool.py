import arxiv
import os
import json
import openpyxl
from openpyxl.utils import get_column_letter
import datetime
import configparser
import time

# read the config file.
config = configparser.ConfigParser()
config.read('config.ini')
# create two dictionary to decode the search sorting settings passing ;from main.
criteria_decode = {
    "--rel": arxiv.SortCriterion.Relevance,
    "--smd": arxiv.SortCriterion.SubmittedDate,
    "--lud": arxiv.SortCriterion.LastUpdatedDate
}

order_decode = {
    "--des": arxiv.SortOrder.Descending,
    "--asc": arxiv.SortOrder.Ascending
}

def cls():
    # function to clear the screen
    if os.name == 'nt':
        _ = os.system('cls')
        # windows
    else:
        _ = os.system('clear')
        # mac or linux

# decode the search max results settings from the config.ini
def max_result(num_download):
    max = config['common_search_options']['max_results']
    if max == 'inf':
        return float('inf')
    elif max.isdigit():
        max = int(max)
        if max == 0:
            max = 10
            # 0 for the default setting
        if num_download > max:
            print(f"The number of download cannot exceed {max}. \nYou should enter a smaller number or set the maximum result in config.ini.")
            exit(1)
        return max

# Workflow: search (get search result) -> download (get download result) -> history (combine information
# from search result and download result)
def arxiv_search(download_path, search_term, num_download, criterion, order):
    search = arxiv.Search(
        query = search_term,
        max_results = max_result(num_download), # tunable in config.ini
        sort_by = criteria_decode[criterion],
        sort_order = order_decode[order]
    )
    result_download(search, num_download, download_path)
    exit(0)
    # exit the program.

## Result download
def result_download(search, num_download, download_path):
# download the pdf file from the search, the number of download is the number of papers we want to download, save in
# download_path. 
    results = search.results()
    # generator of the search result.
    downloaded_papers = []
    # list of paper object that we have downloaded.
    download_count = 0
    # count the number of papers we have downloaded.
    saved_paper_id = []
    if os.path.exists("history.json"):
        # check if the history file exists.
        with open('history.json', 'r') as f:
            for line in f:
                history = json.loads(line)
                for paper in history['Downloaded Result']:
                    saved_paper_id.append(paper['id'])
                    # append the paper object to the list of downloaded papers.

    while download_count < num_download:
        try:
            paper = next(results)
            if paper.get_short_id() in saved_paper_id:
                # check if the paper has been downloaded.
                continue
            paper.download_pdf(dirpath=download_path, filename=f"{paper.get_short_id()}.pdf")
            # download the pdf file. name is the paper id.
        except StopIteration:
            # stop the generator when no more results.
            print("No more results.")
            break
            # generator end.
        except arxiv.exceptions.HTTPError:
            # handle the error when the paper is not available.
            print(f"Paper [{paper.title}] is not available.")
            continue
        except Exception as e:
            print(f"Error occurred while downloading paper [{paper.title}]: {str(e)}")
            continue
        else:
            download_count += 1
            downloaded_papers.append(paper)
            # append the paper object to the list of downloaded papers.

    save_result(search, downloaded_papers)


## Result save
def list_authors(authors):
    # convert the list of authors object to a string with authors.
    str_authors = []
    if authors == []:
        return "No author"
    else:
        for author in authors:
            str_authors.append(author.name)
        return ", ".join(str_authors)

def save_result(search, download_result):
# save the search result in json file.
    results = {
        "Search Term": search.query,
        "Date": datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "Items downloaded": len(download_result),
        "Downloaded Result": []
    }
    if config['download_history'].getboolean('record_search_term') == True:
        results["Items found"] = len(list(search.results()))
        # record the number of items found. (Extremely time costly, in default disabled)
    for result in download_result:
        # Extract relevant information from the search result
        result_info = {
            'title': result.title,
            'authors': list_authors(result.authors),
            'abstract': result.summary,
            'upload_date': result.updated.strftime("%Y-%m-%d %H:%M:%S"),
            'doi': result.doi,
            'categories': result.categories,
            'id': result.get_short_id(),
            'pdf_url': result.pdf_url,
            # Add more information as needed
        }
        results['Downloaded Result'].append(result_info)

    # then save this dictionary in json file or excel file if needed.
    with open('history.json', 'a') as json_file:
        json.dump(results, json_file)
        # no indent to make each root a single line.
        json_file.write('\n') #start a new line

# save in excel 
    if config['download_history'].getboolean('save_excel') == True:
        save_result_excel(results)

def save_result_excel(results):
    if os.path.exists("history.xlsx"):
        # check if the history file exists.
        # if it exists, then create a new sheet.
        # if it doesn't exist, then create a new workbook.
        # the sheet name will be the search term with date.
        wb = openpyxl.load_workbook("history.xlsx")
        ws = wb.create_sheet(f"Search Result-{results['Search Term']}-{results['Date']}")
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Search Result-{results['Search Term']}-{results['Date']}"

    ws.append(['Title', 'Authors', 'Abstract', 'Upload Date', 'DOI', 'Categories'])
    # create the header of the excel file. [details of the downloaded articles]
    for result in results['Downloaded Result']:
        ws.append([result['title'], result['authors'], result['abstract'], result['upload_date'], result['doi'], ", ".join(result['categories'])])
    
    ws.append([''])
    # data about the search
    if 'Items found' in results:
        ws.append(['Search Term', 'Date', 'Items found', 'Items downloaded'])
        ws.append([results['Search Term'], results['Date'], results['Items found'], results['Items downloaded']])
    else:
        ws.append(['Search Term', 'Date', 'Items downloaded'])
        ws.append([results['Search Term'], results['Date'], results['Items downloaded']])
    # save the file.
    wb.save(os.path.join(os.getcwd(), f'history.xlsx'))

def print_record(book, page, current_page):
    if page == 0 or page == current_page - 1:
        for record in book[(current_page-1)*10:]:
            print(record)
    else:
        for record in book[(current_page-1)*10:(current_page)*10]:
            print(record)

def print_legend(page, current_page):
    cls()
    if page == 0:
        print(f"Page {current_page} of {page + 1}")
        print("[e] Exit\n")
        print('-' * 20)
    else:
        print(f"Page {current_page} of {page + 1}")
        print("[p] Previous Page [n] Next Page\n[e] Exit\n")
        print('-' * 20)

def get_idx(book, page, current_page):
    if page == 0:
        print_legend(page, current_page)
        print_record(book, page, current_page)
        idx = input("Please enter the record number:\n").lower()
        while not idx.isdigit():
            if idx == 'e':
                exit()
            print("Invalid input, please enter a number.")
            idx = input("Please enter the record number:")
        return int(idx) - 1
    else:
        print_legend(page, current_page)
        print_record(book, page, current_page)
        idx = input("Please enter the record number or instruction:\n").lower()
        while not idx.isdigit() and idx != 'p' and idx != 'n' and idx != 'e':
            print("Invalid input, please enter a number or instruction.")
            idx = input("Please enter the record number or instruction:\n")
            # check if the input is a number or instruction.
        # instruction
        if idx == 'e':
            exit()
        elif idx == 'p':
            if current_page == 1:
                print("This is the first page.")
                time.sleep(1)
                return get_idx(book, page, current_page)
            else:
                print_record(book, page, current_page - 1)
                return get_idx(book, page, current_page - 1)
        elif idx == 'n':
            if current_page == page + 1:
                print("This is the last page.")
                time.sleep(1)
                return get_idx(book, page, current_page)
            else:
                print_record(book, page, current_page + 1)
                return get_idx(book, page, current_page + 1)
        
        idx = int(idx)
        if (current_page - 1) * 10 < idx <= (current_page) * 10:
            # e.g. idx = 10, the first page shows 1-10
            return idx - 1
        else:
            print("Invalid input, please enter a number IN THE SCREEN.")
            time.sleep(1)
            return get_idx(book, page, current_page)

def page_choose(book):
    page = len(book) // 10
    # number of pages, must >=0
    if page == 0:
        idx = get_idx(book, page, 1)
        return idx
    else: 
        # more than 10 lines
        idx = get_idx(book, page, 1)
        return idx
        
def show_history():
    # show the history of the search.
    # In formatted way
    if os.path.exists("history.json"):
        with open('history.json', 'r') as f:
            # Implenment a 'page', create a buffer to store the information
            # each page will have 10 lines.
            book = []
            # store the history in a list of strings.
            book_data = []
            # store the history in a list of dictionaries.
            count = 0
            for line in f:
                # append all the history as a 'book'
                count += 1
                history = json.loads(line)
                if 'Items found' in history:
                    msg = f"{count}. Search Term: {history['Search Term']}, Date: {history['Date']}, Items found: {history['Items found']}, Items downloaded: {history['Items downloaded']}\n"
                else:  
                    msg = f"{count}. Search Term: {history['Search Term']}, Date: {history['Date']}, Items downloaded: {history['Items downloaded']}\n"
                book.append(msg)
                book_data.append(history)

        index = page_choose(book)
        # choose the record to show.
        history_to_show = book_data[index]
        cls()
        for item in history_to_show['Downloaded Result']:
            print(f"Title: {item['title']}\nAuthors: {item['authors']}\nAbstract: {item['abstract']}\nUpload Date: {item['upload_date']}\nDOI: {item['doi']}\nCategories: {item['categories']}\n")
            print("="*100)
        exit(0)                
    else:
        cls()
        print("No history yet.")
        input("Press any key to exit.")
        exit(0)